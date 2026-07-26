#!/usr/bin/env python3
"""Scrape Steve Silver game/dining/server products → Volusion import CSV + images."""
from __future__ import annotations

import csv
import json
import os
import re
import time
import urllib.error
import urllib.request
from html import unescape
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "docs" / "imports" / "steve-silver-volusion"
PHOTO_DIR = ROOT / "vspfiles" / "photos"
URLS_PATH = Path("/tmp/ss-import/target-urls.json")
UA = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml",
}


def get(url: str, retries: int = 6) -> tuple[str, str]:
    for i in range(retries):
        try:
            req = urllib.request.Request(url, headers=UA)
            with urllib.request.urlopen(req, timeout=55) as r:
                return r.read().decode("utf-8", "replace"), r.geturl()
        except Exception:
            time.sleep(1.5 + i * 2)
    return "", url


def clean_text(s: str) -> str:
    s = unescape(s or "")
    s = re.sub(r"<br\s*/?>", "\n", s, flags=re.I)
    s = re.sub(r"</p\s*>", "\n\n", s, flags=re.I)
    s = re.sub(r"<[^>]+>", " ", s)
    s = s.replace("\xa0", " ")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def volusion_code(sku: str) -> str:
    sku = (sku or "").strip().upper()
    sku = sku.replace(" ", "-")
    sku = re.sub(r"[^A-Z0-9_-]", "-", sku)
    sku = re.sub(r"-{2,}", "-", sku).strip("-")
    if not sku.startswith("SS-"):
        sku = "SS-" + sku
    return sku


def extract_images(html: str, vendor_sku: str = "") -> list[str]:
    """Pull gallery images for this product only (SKU in filename).

    Related-product / "you may also like" images on the page are ignored when
    a vendor SKU is known, so e.g. Colvin Green does not get Black server shots.
    """
    # Prefer Woo gallery payloads over scraping every upload URL on the page.
    raw: list[str] = []
    for pat in (
        r'data-large_image="([^"]+)"',
        r'woocommerce-product-gallery__image[\s\S]{0,500}?href="([^"]+)"',
        r"(https://stevesilver\.com/wp-content/uploads/\d{4}/\d{2}/[^\"'\s]+\.(?:jpe?g|png|webp))",
    ):
        raw.extend(re.findall(pat, html, re.I))

    aliases = []
    sku = (vendor_sku or "").strip().upper()
    if sku:
        aliases.append(sku)
        # Ivory Colvin gallery files use COL500LWSV while page SKU is COL500WSV.
        if sku == "COL500WSV":
            aliases.append("COL500LWSV")
    aliases_l = [a.lower() for a in aliases]

    out: list[str] = []
    seen: set[str] = set()
    for u in raw:
        if not u.startswith("http"):
            continue
        u = u.split("?")[0]
        # skip resized derivatives: name-100x100.jpg etc.
        if re.search(r"-\d+x\d+\.(?:jpe?g|png|webp)$", u, re.I):
            continue
        if any(x in u.lower() for x in ("logo", "favicon", "sprite", "placeholder")):
            continue
        base = u.rsplit("/", 1)[-1]
        bl = base.lower()
        if base in seen:
            continue
        if aliases_l and not any(a in bl for a in aliases_l):
            continue
        if "swatch" in bl:
            continue
        seen.add(base)
        out.append(u)

    # Product studio shots (WS) first; deprioritize multi-SKU room sets.
    def rank(u: str) -> tuple:
        b = u.lower()
        score = 0
        if re.search(r"_ws1\b|_ws1\.", b):
            score -= 100
        elif re.search(r"_ws2a?\b|_ws2a?\.", b):
            score -= 80
        elif re.search(r"_ws\d", b):
            score -= 70
        elif re.search(r"_vg\d|_dtl", b):
            score -= 50
        if re.search(r"_rs\d|morefloor", b):
            score += 80
        if "dims" in b or "_dim" in b:
            score += 120
        if "chair" in b and (not aliases_l or not any(a in b for a in aliases_l)):
            score += 100
        return (score, u)

    return sorted(out, key=rank)


def extract_product(url: str, html: str) -> dict:
    title_m = re.search(
        r'<h1[^>]*class="[^"]*product_title[^"]*"[^>]*>([^<]+)', html, re.I
    )
    name = clean_text(title_m.group(1) if title_m else "")
    name = re.sub(r"\s*\[product_price[^\]]*\]\s*", "", name).strip()

    sku = ""
    for pat in [
        r'"sku"\s*:\s*"([^"]+)"',
        r'data-product_sku="([^"]+)"',
        r'<span class="sku">([^<]+)',
    ]:
        m = re.search(pat, html, re.I)
        if m:
            sku = m.group(1).strip()
            break

    # Short description = productdescription
    short = ""
    m = re.search(
        r'class="[^"]*woocommerce-product-details__short-description[^"]*"[^>]*>([\s\S]*?)</div>',
        html,
        re.I,
    )
    if m:
        short = clean_text(m.group(1))
    if not short:
        # JSON-LD Product description
        for block in re.findall(
            r'<script type="application/ld\+json"[^>]*>([\s\S]*?)</script>', html, re.I
        ):
            try:
                data = json.loads(block)
            except Exception:
                continue
            nodes = data.get("@graph", [data]) if isinstance(data, dict) else data
            if not isinstance(nodes, list):
                nodes = [nodes]
            for n in nodes:
                if isinstance(n, dict) and n.get("@type") == "Product" and n.get("description"):
                    short = clean_text(n["description"])
                    break

    # Tech specs = description tab bullets starting at "Set includes" / feature list
    bullets: list[str] = []
    # Prefer tab-description list items
    tab = re.search(
        r'id="tab-description"([\s\S]*?)(?:id="tab-|</div>\s*</div>\s*<div class=")',
        html,
        re.I,
    )
    scope = tab.group(1) if tab else html
    for li in re.findall(r"<li[^>]*>([\s\S]*?)</li>", scope, re.I):
        t = clean_text(li)
        if len(t) < 12:
            continue
        if t.startswith("{") or "application/ld+json" in t:
            continue
        if "Steve Silver Company" in t and len(t) > 200:
            continue
        bullets.append(t)

    # Keep from first "Set includes" / "Crafted of" / "Optional Server" onward when present
    start = 0
    for i, b in enumerate(bullets):
        if re.match(r"(?i)(set includes|crafted of|optional server)", b):
            start = i
            break
    tech = "\n".join(f"• {b}" for b in bullets[start:])

    # Price
    price = ""
    for pat in [
        r'"price"\s*:\s*"([0-9.]+)"',
        r'"display_price"\s*:\s*([0-9.]+)',
        r'woocommerce-Price-amount[^>]*>\s*<bdi><span[^>]*>[^<]*</span>([0-9,.]+)',
        r'itemprop="price"[^>]*content="([0-9.]+)"',
    ]:
        m = re.search(pat, html, re.I)
        if m:
            price = m.group(1).replace(",", "")
            break

    # Weight
    weight = ""
    for pat in [
        r'"weight"\s*:\s*"([^"]+)"',
        r">Weight</[^>]*>\s*<[^>]+>([^<]+)",
        r"Product Weight[:\s]*</[^>]+>\s*<[^>]+>([^<]+)",
    ]:
        m = re.search(pat, html, re.I)
        if m:
            weight = clean_text(m.group(1))
            # normalize to pounds number if possible
            num = re.search(r"([0-9]+(?:\.[0-9]+)?)", weight)
            if num:
                weight = num.group(1)
            break

    # Availability
    availability = "Available"
    stock = re.search(r"(\d+)\s+in stock", html, re.I)
    if re.search(r"temporarily unavailable|out of stock", html, re.I):
        availability = "Out of Stock"
    elif stock:
        availability = "Available"

    images = extract_images(html, vendor_sku=sku)
    group = "other"
    low = (name + " " + url).lower()
    if "game" in low or "tournament" in low or "cambridge" in low and "game" in low or "rylie" in low:
        group = "game"
    elif "server" in low or "curio" in low or "sideboard" in low:
        group = "server"
    else:
        group = "dining"

    return {
        "source_url": url,
        "vendor_sku": sku,
        "productcode": volusion_code(sku) if sku else "",
        "productname": name,
        "productprice": price,
        "productweight": weight,
        "freeshipping": "N",
        "availability": availability,
        "productdescription": short,
        "techspecs": tech,
        "image_urls": images,
        "group": group,
    }


def download_volusion_images(code: str, urls: list[str]) -> list[str]:
    """Save as CODE-1.jpg + CODE-altviewN.jpg (Volusion convention)."""
    PHOTO_DIR.mkdir(parents=True, exist_ok=True)
    saved: list[str] = []
    if not code or not urls:
        return saved
    # Cap altviews
    urls = urls[:16]
    for idx, url in enumerate(urls):
        if idx == 0:
            fname = f"{code}-1.jpg"
        else:
            fname = f"{code}-altview{idx}.jpg"
        dest = PHOTO_DIR / fname
        for attempt in range(5):
            try:
                req = urllib.request.Request(url, headers=UA)
                with urllib.request.urlopen(req, timeout=60) as r:
                    data = r.read()
                if len(data) < 500:
                    raise ValueError("tiny")
                dest.write_bytes(data)
                # also write -2T thumbnail proxy from main for Volusion PLP habits
                if idx == 0:
                    tdest = PHOTO_DIR / f"{code}-2T.jpg"
                    if not tdest.exists():
                        tdest.write_bytes(data)
                saved.append(fname)
                break
            except Exception:
                time.sleep(1 + attempt)
        time.sleep(0.35)
    return saved


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    payload = json.loads(URLS_PATH.read_text())
    urls = payload["all"]
    products: list[dict] = []
    for i, url in enumerate(urls, 1):
        print(f"[{i}/{len(urls)}] {url}")
        html, final = get(url)
        if not html:
            print("  FAIL fetch")
            continue
        p = extract_product(final, html)
        if not p["productcode"]:
            # fallback from slug
            slug = final.rstrip("/").split("/")[-1]
            p["productcode"] = volusion_code(slug[:40])
            p["vendor_sku"] = slug
        print(" ", p["productcode"], p["productname"][:60], "imgs", len(p["image_urls"]))
        imgs = download_volusion_images(p["productcode"], p["image_urls"])
        p["volusion_images"] = ";".join(imgs)
        products.append(p)
        time.sleep(0.8)

    # Existing site server audit notes
    existing_servers = [
        {
            "productcode": "SS-BUR500NSV",
            "productname": "Magnolia Cathedral Doored Server, Black",
            "on_site": "Y",
            "notes": (
                "Already on site (cat 217). Primary hero often uses -2T; ensure "
                "SS-BUR500NSV-1.jpg is present. Description OK but should use "
                "stevesilver.com short description + techspecs bullet format for import consistency."
            ),
        },
        {
            "productcode": "SS-AUB500SV",
            "productname": "Auburn Server (listed; Aubrey group)",
            "on_site": "Y",
            "notes": (
                "Already on site as SS-AUB500SV but copy looks wrong (mentions expanding table). "
                "Replace productdescription/techspecs from Aubrey Server Black page; "
                "upload missing -1.jpg (altviews exist). Confirm naming: Auburn vs Aubrey."
            ),
        },
        {
            "productcode": "KARINA-SIDEBOARD",
            "productname": "Karina Sideboard",
            "on_site": "Y",
            "notes": "Already on site. Prefer SS- style code if re-importing; images present.",
        },
    ]

    cols = [
        "productcode",
        "productname",
        "productprice",
        "productweight",
        "freeshipping",
        "availability",
        "productdescription",
        "techspecs",
        "group",
        "vendor_sku",
        "source_url",
        "volusion_images",
    ]
    csv_path = OUT_DIR / "steve_silver_game_dining_server_import.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for p in products:
            w.writerow(p)

    # Game-only sheet convenience
    game_path = OUT_DIR / "steve_silver_game_table_sets_import.csv"
    with game_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for p in products:
            if p["group"] == "game":
                w.writerow(p)

    dining_path = OUT_DIR / "steve_silver_dining_groups_import.csv"
    with dining_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for p in products:
            if p["group"] in ("dining", "server"):
                w.writerow(p)

    servers_path = OUT_DIR / "servers_on_site_format_notes.csv"
    with servers_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f, fieldnames=["productcode", "productname", "on_site", "notes"]
        )
        w.writeheader()
        w.writerows(existing_servers)
        for p in products:
            if p["group"] == "server":
                w.writerow(
                    {
                        "productcode": p["productcode"],
                        "productname": p["productname"],
                        "on_site": "N",
                        "notes": "New from Steve Silver scrape; images saved as Volusion -1 / -altviewN.",
                    }
                )

    # Try xlsx via openpyxl if available; else csv is enough
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "all_products"
        ws.append(cols)
        for p in products:
            ws.append([p.get(c, "") for c in cols])
        ws2 = wb.create_sheet("game_sets")
        ws2.append(cols)
        for p in products:
            if p["group"] == "game":
                ws2.append([p.get(c, "") for c in cols])
        ws3 = wb.create_sheet("dining_and_servers")
        ws3.append(cols)
        for p in products:
            if p["group"] in ("dining", "server"):
                ws3.append([p.get(c, "") for c in cols])
        ws4 = wb.create_sheet("servers_site_notes")
        ws4.append(["productcode", "productname", "on_site", "notes"])
        for row in existing_servers:
            ws4.append([row[k] for k in ["productcode", "productname", "on_site", "notes"]])
        xlsx_path = OUT_DIR / "steve_silver_game_dining_server_import.xlsx"
        wb.save(xlsx_path)
        print("wrote", xlsx_path)
    except Exception as e:
        print("xlsx skipped:", e)

    summary = {
        "product_count": len(products),
        "game": sum(1 for p in products if p["group"] == "game"),
        "dining": sum(1 for p in products if p["group"] == "dining"),
        "server": sum(1 for p in products if p["group"] == "server"),
        "csv": str(csv_path.relative_to(ROOT)),
    }
    (OUT_DIR / "summary.json").write_text(json.dumps(summary, indent=2))
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
