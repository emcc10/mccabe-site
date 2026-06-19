#!/usr/bin/env python3
"""Build Steve Silver bed catalog spreadsheet and Volusion-ready images.

Matches bedroom collections that already have nightstand, dresser, and chest SKUs.
Fetches king + queen bed pages from stevesilver.com, downloads the best piece-only
hero image, writes -1.jpg (main) and -1T.jpg (thumbnail), and exports a CSV spreadsheet.
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import re
import shutil
import sys
import urllib.error
import urllib.request
from html import unescape
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "catalog" / "steve-silver-beds"
IMAGE_DIR = OUT_DIR / "images"
VOLUSION_PHOTOS = ROOT / "vspfiles" / "photos"
CSV_PATH = OUT_DIR / "steve_silver_beds.csv"
UA = {"User-Agent": "Mozilla/5.0 (McCabe Steve Silver bed catalog)"}
MIN_BYTES = 15_000
THUMB_MAX = (900, 700)
JPEG_QUALITY = 90

SKIP_MARKERS = ("LOGO", "INFO", "HARDWARE", "AMP1", "MOBILE.PNG", "COVER")
ROOM_MARKERS = ("_RS", "REVISED", "BRREG", "K4PC", "K5PC", "KS4PC", "-K4PC", "-K5PC", "BEDROOM", "_LS")
PIECE_MARKERS = ("_WS", "_VG1", "_VG2", "KFB", "QFB", "KHB", "QHB", "KBED", "QBED", "STORAGE", "KSBED", "QSBED")
BED_MARKERS = ("KFB", "QFB", "KHB", "QHB", "KBED", "QBED", "KSBED", "QSBED", "CKR", "SRB")
OTHER_PIECE_MARKERS = ("900NS", "900DR", "900CT", "900MR", "900NSS", "900DRS", "NIGHTSTAND", "DRESSER", "CHEST", "MIRROR")

# Collections aligned with existing SS-BC900/BC950, CAS900, HP900, MON900, RV900, SIG900 pieces.
BED_PRODUCTS: list[dict[str, str]] = [
    {
        "volusion_code": "SS-BC900KFB",
        "collection": "Bear Creek",
        "finish": "White Smoke / Honey Smoke",
        "size": "King",
        "page": "https://stevesilver.com/product/bear-creek-3-piece-king-bed/",
        "sku_hint": "BC900KFB",
        "related": "SS-BC900NS, SS-BC900DR, SS-BC900CTT, SS-BC900MR",
        "extra_pages": "https://stevesilver.com/product/bear-creek-queen-headboard/",
    },
    {
        "volusion_code": "SS-BC900QFB",
        "collection": "Bear Creek",
        "finish": "White Smoke / Honey Smoke",
        "size": "Queen",
        "page": "https://stevesilver.com/product/bear-creek-3-piece-queen-bed/",
        "sku_hint": "BC900QFB",
        "related": "SS-BC900NS, SS-BC900DR, SS-BC900CTT, SS-BC900MR",
        "extra_pages": "https://stevesilver.com/product/bear-creek-queen-headboard/",
    },
    {
        "volusion_code": "SS-BC950KFB",
        "collection": "Bear Creek",
        "finish": "Brown",
        "size": "King",
        "page": "https://stevesilver.com/product/bear-creek-king-bed-brown/",
        "sku_hint": "BC950KFB",
        "related": "SS-BC950NSB, SS-BC950DRB, SS-BC950CTBT, SS-BC950MRB",
    },
    {
        "volusion_code": "SS-BC950QFB",
        "collection": "Bear Creek",
        "finish": "Brown",
        "size": "Queen",
        "page": "https://stevesilver.com/product/bear-creek-queen-bed/",
        "sku_hint": "BC950QFB",
        "related": "SS-BC950NSB, SS-BC950DRB, SS-BC950CTBT, SS-BC950MRB",
        "extra_pages": "https://stevesilver.com/product/bear-creek-king-bed-brown/",
    },
    {
        "volusion_code": "SS-CAS900KFB",
        "collection": "Cassie Illuminating",
        "finish": "Shimmering Pearl",
        "size": "King",
        "page": "https://stevesilver.com/product/cassie-illuminating-king-bed/",
        "sku_hint": "CAS900KFB",
        "related": "SS-CAS900NS, SS-CAS900DR, SS-CAS900C, SS-CAS900M",
    },
    {
        "volusion_code": "SS-CAS900QFB",
        "collection": "Cassie Illuminating",
        "finish": "Shimmering Pearl",
        "size": "Queen",
        "page": "https://stevesilver.com/product/cassie-illuminating-queen-bed/",
        "sku_hint": "CAS900QFB",
        "related": "SS-CAS900NS, SS-CAS900DR, SS-CAS900C, SS-CAS900M",
    },
    {
        "volusion_code": "SS-HP900KFBW",
        "collection": "Highland Park",
        "finish": "Cathedral White",
        "size": "King",
        "page": "https://stevesilver.com/product/highland-park-3-piece-king-bed-cathedral-white/",
        "sku_hint": "HP900KFBW",
        "related": "SS-HP900NSW, SS-HP900CTWT, SS-HP900MRW",
        "image_override": "https://stevesilver.com/wp-content/uploads/2019/08/HP900QFBW.jpg",
    },
    {
        "volusion_code": "SS-HP900QFBW",
        "collection": "Highland Park",
        "finish": "Cathedral White",
        "size": "Queen",
        "page": "https://stevesilver.com/product/highland-park-3-piece-queen-bed-cathedral-white/",
        "sku_hint": "HP900QFBW",
        "related": "SS-HP900NSW, SS-HP900CTWT, SS-HP900MRW",
        "image_override": "https://stevesilver.com/wp-content/uploads/2019/08/HP900QFBW.jpg",
    },
    {
        "volusion_code": "SS-HP900KFBD",
        "collection": "Highland Park",
        "finish": "Waxed Driftwood",
        "size": "King",
        "page": "https://stevesilver.com/product/highland-park-3-piece-king-bed-waxed-driftwood/",
        "sku_hint": "HP900KFBD",
        "related": "SS-HP900NSD, SS-HP900CTDT, SS-HP900MRD",
    },
    {
        "volusion_code": "SS-HP900QFBD",
        "collection": "Highland Park",
        "finish": "Waxed Driftwood",
        "size": "Queen",
        "page": "https://stevesilver.com/product/highland-park-3-piece-queen-bed-waxed-driftwood/",
        "sku_hint": "HP900QFBD",
        "related": "SS-HP900NSD, SS-HP900CTDT, SS-HP900MRD",
    },
    {
        "volusion_code": "SS-MON900KFB",
        "collection": "Montana",
        "finish": "Sand",
        "size": "King",
        "page": "https://stevesilver.com/product/montana-king-bed-sand/",
        "sku_hint": "MON900KFB",
        "related": "SS-MON900NSS, SS-MON900DRST, SS-MON900CS",
    },
    {
        "volusion_code": "SS-MON900QFB",
        "collection": "Montana",
        "finish": "Sand",
        "size": "Queen",
        "page": "https://stevesilver.com/product/montana-queen-bed-sand/",
        "sku_hint": "MON900QFB",
        "related": "SS-MON900NSS, SS-MON900DRST, SS-MON900CS",
    },
    {
        "volusion_code": "SS-RV900KFB",
        "collection": "Riverdale",
        "finish": "Natural",
        "size": "King",
        "page": "https://stevesilver.com/product/riverdale-king-storage-bed/",
        "sku_hint": "RV900KFB",
        "related": "SS-RV900NS, SS-RV900DR, SS-RV900C, SS-RV900M",
    },
    {
        "volusion_code": "SS-RV900QFB",
        "collection": "Riverdale",
        "finish": "Natural",
        "size": "Queen",
        "page": "https://stevesilver.com/product/riverdale-queen-storage-bed/",
        "sku_hint": "RV900QFB",
        "related": "SS-RV900NS, SS-RV900DR, SS-RV900C, SS-RV900M",
    },
    {
        "volusion_code": "SS-SIG900KFB",
        "collection": "Sigmund",
        "finish": "Gray",
        "size": "King",
        "page": "https://stevesilver.com/product/sigmund-king-bed/",
        "sku_hint": "SIG900KFB",
        "related": "SS-SIG900NS, SS-SIG900DR, SS-SIG900C",
    },
    {
        "volusion_code": "SS-SIG900QFB",
        "collection": "Sigmund",
        "finish": "Gray",
        "size": "Queen",
        "page": "https://stevesilver.com/product/sigmund-queen-bed/",
        "sku_hint": "SIG900QFB",
        "related": "SS-SIG900NS, SS-SIG900DR, SS-SIG900C",
    },
]


def fetch_text(url: str, retries: int = 4) -> str:
    last_exc: Exception | None = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=UA)
            with urllib.request.urlopen(req, timeout=45) as resp:
                return resp.read().decode("utf-8", "replace")
        except urllib.error.HTTPError as exc:
            last_exc = exc
            if exc.code in (429, 503) and attempt + 1 < retries:
                import time  # noqa: PLC0415

                time.sleep(2 * (attempt + 1))
                continue
            raise
    if last_exc:
        raise last_exc
    raise RuntimeError(f"fetch failed: {url}")


def fetch_bytes(url: str) -> bytes:
    req = urllib.request.Request(url, headers=UA)
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = resp.read()
    if len(data) < MIN_BYTES:
        raise ValueError(f"image too small ({len(data)} bytes): {url}")
    return data


def strip_html(text: str) -> str:
    text = unescape(re.sub(r"<[^>]+>", " ", text))
    return " ".join(text.split())


def parse_json_ld_product(html: str) -> dict[str, str]:
    out: dict[str, str] = {}
    for block in re.findall(r'<script type="application/ld\+json"[^>]*>(.*?)</script>', html, re.I | re.S):
        try:
            data = json.loads(block)
        except json.JSONDecodeError:
            continue
        items = data.get("@graph", [data]) if isinstance(data, dict) else data
        if not isinstance(items, list):
            items = [items]
        for item in items:
            if isinstance(item, dict) and item.get("@type") == "Product":
                out["product_name"] = str(item.get("name") or "").strip()
                out["description"] = strip_html(str(item.get("description") or ""))
                out["steve_silver_sku"] = str(item.get("sku") or "").strip()
                break
        if out:
            break
    return out


def parse_bullets(html: str) -> list[str]:
    m = re.search(r'id="tab-description"[^>]*>(.*?)</div>\s*<div', html, re.I | re.S)
    if not m:
        return []
    bullets: list[str] = []
    for li in re.findall(r"<li[^>]*>(.*?)</li>", m.group(1), re.I | re.S):
        text = strip_html(li)
        if text and text.lower() != "description":
            bullets.append(text)
    return bullets


def parse_attributes(html: str) -> list[str]:
    m = re.search(
        r'Additional information.*?<table class="woocommerce-product-attributes[^"]*"[^>]*>(.*?)</table>',
        html,
        re.I | re.S,
    )
    if not m:
        return []
    lines: list[str] = []
    for row in re.findall(r"<tr[^>]*>(.*?)</tr>", m.group(1), re.I | re.S):
        th = re.search(r"<th[^>]*>(.*?)</th>", row, re.I | re.S)
        td = re.search(r"<td[^>]*>(.*?)</td>", row, re.I | re.S)
        if th and td:
            k = strip_html(th.group(1))
            v = strip_html(td.group(1))
            if k and v:
                lines.append(f"{k}: {v}")
    return lines


def build_techspecs(description: str, bullets: list[str], attrs: list[str]) -> str:
    parts: list[str] = []
    if description:
        parts.append(description)
    if bullets:
        parts.append("")
        parts.append("Features:")
        parts.extend(f"• {b}" for b in bullets)
    if attrs:
        parts.append("")
        parts.append("Specifications:")
        parts.extend(f"• {a}" for a in attrs)
    return "\n".join(parts).strip()


def extract_gallery_images(html: str) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    seen: set[str] = set()
    for pat in (
        r'data-large_image="([^"]+)"',
        r'data-src="(https://[^"]+\.(?:jpg|jpeg|png|webp))"',
        r'src="(https://stevesilver\.com/wp-content/uploads/[^"]+\.(?:jpg|jpeg|png|webp))"',
    ):
        for m in re.finditer(pat, html, re.I):
            url = m.group(1).replace("\\/", "/")
            if url in seen or re.search(r"-\d+x\d+\.", url, re.I):
                continue
            seen.add(url)
            alt = url.rsplit("/", 1)[-1]
            text = (alt + " " + url).upper()
            if any(x in text for x in SKIP_MARKERS):
                continue
            out.append((alt, url))
    return out


def is_room_scene(alt: str, url: str) -> bool:
    text = (alt + " " + url).upper()
    if "_LS" in text and not any(x in text for x in BED_MARKERS):
        return True
    return any(marker in text for marker in ROOM_MARKERS if marker != "_LS")


def is_bed_detail(alt: str, url: str) -> bool:
    text = (alt + " " + url).upper()
    if "DTL" not in text:
        return False
    return any(x in text for x in BED_MARKERS)


def is_combo_shot(text: str) -> bool:
    other_markers = ("900NS", "900DR", "900MR", "900CT", "DRB", "MRB", "NSS", "CTB", "CTBT", "NSB", "_DR_", "_MR_", "_NS_", "_CT_")
    bed_markers = BED_MARKERS + ("KFBB", "KHBB", "QFBB", "QHBB", "SRB")
    has_bed = any(x in text for x in bed_markers)
    other_hits = sum(1 for x in other_markers if x in text)
    if other_hits >= 1 and has_bed and any(x in text for x in other_markers):
        return True
    if other_hits >= 2:
        return True
    if any(x in text for x in ("CTB", "CTBT", "DRB", "MRB", "NSB")) and not has_bed:
        return True
    return False


def should_skip_image(alt: str, url: str) -> bool:
    text = (alt + " " + url).upper()
    if any(marker in text for marker in SKIP_MARKERS):
        return True
    if is_room_scene(alt, url):
        return True
    if is_combo_shot(text):
        return True
    if any(x in text for x in OTHER_PIECE_MARKERS) and not any(x in text for x in BED_MARKERS):
        return True
    return False


def size_match_score(text: str, size: str) -> int:
    score = 0
    if size == "King":
        if any(x in text for x in ("KFB", "KHB", "KBED", "KSBED")):
            score += 80
        if any(x in text for x in ("QFB", "QHB", "QBED", "QSBED")):
            score -= 100
    if size == "Queen":
        if any(x in text for x in ("QFB", "QHB", "QBED", "QSBED")):
            score += 80
        if any(x in text for x in ("KFB", "KHB", "KBED", "KSBED")) and "Q" not in text:
            score -= 40
    return score


def piece_score(alt: str, url: str, size: str, sku_hint: str, *, detail: bool = False) -> int:
    text = (alt + " " + url).upper()
    if should_skip_image(alt, url) and not detail:
        return -999
    if detail and not is_bed_detail(alt, url):
        return -999
    score = 0
    if "_WS" in text:
        score += 100
    if "_VG1" in text:
        score += 90
    score += size_match_score(text, size)
    if sku_hint.upper() in text:
        score += 40
    prefix = re.sub(r"(KFB|QFB|KHB|QHB|KBED|QBED).*$", "", sku_hint.upper())
    if prefix and prefix in text:
        score += 30
    if any(x in text for x in BED_MARKERS):
        score += 25
    if any(x in text for x in ("KFBB", "KHBB", "QFBB", "QHBB", "SRB")):
        score += 70
    if detail:
        score += 10
    return score


def pick_image_url(
    images: list[tuple[str, str]],
    size: str,
    sku_hint: str,
    image_override: str = "",
) -> str:
    for detail in (False, True):
        candidates = [
            (piece_score(alt, url, size, sku_hint, detail=detail), url)
            for alt, url in images
        ]
        candidates = [c for c in candidates if c[0] > 0]
        candidates.sort(key=lambda x: x[0], reverse=True)
        if candidates:
            return candidates[0][1]
    if image_override:
        return image_override
    for alt, url in images:
        if not should_skip_image(alt, url):
            return url
    return images[0][1] if images else ""


def to_jpeg(data: bytes) -> bytes:
    from PIL import Image  # noqa: PLC0415

    img = Image.open(io.BytesIO(data))
    if img.mode not in ("RGB", "L"):
        img = img.convert("RGB")
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=JPEG_QUALITY, optimize=True)
    return buf.getvalue()


def make_thumbnail(src: Path, dest: Path) -> int:
    from PIL import Image  # noqa: PLC0415

    img = Image.open(src)
    if img.mode not in ("RGB", "L"):
        img = img.convert("RGB")
    w, h = img.size
    max_w, max_h = THUMB_MAX
    scale = min(max_w / w, max_h / h, 1.0)
    new_w = max(1, int(round(w * scale)))
    new_h = max(1, int(round(h * scale)))
    if (new_w, new_h) != (w, h):
        resample = Image.Resampling.LANCZOS if hasattr(Image, "Resampling") else Image.LANCZOS
        img = img.resize((new_w, new_h), resample)
    dest.parent.mkdir(parents=True, exist_ok=True)
    img.save(dest, format="JPEG", quality=JPEG_QUALITY, optimize=True)
    return dest.stat().st_size


def process_product(config: dict[str, str], force: bool) -> dict[str, str]:
    code = config["volusion_code"]
    main_path = IMAGE_DIR / f"{code}-1.jpg"
    thumb_path = IMAGE_DIR / f"{code}-1T.jpg"
    page = config["page"]

    row: dict[str, str] = {
        "volusion_code": code,
        "collection": config["collection"],
        "finish": config["finish"],
        "size": config["size"],
        "steve_silver_url": page,
        "related_pieces": config["related"],
        "product_name": "",
        "steve_silver_sku": "",
        "description": "",
        "techspecs": "",
        "image_main": "",
        "image_thumbnail": "",
        "image_source_url": "",
    }

    print(f"=== {code} ===")
    html = fetch_text(page)
    meta = parse_json_ld_product(html)
    bullets = parse_bullets(html)
    attrs = parse_attributes(html)
    row.update(meta)
    row["techspecs"] = build_techspecs(row.get("description", ""), bullets, attrs)

    images = extract_gallery_images(html)
    for extra in config.get("extra_pages", "").split(","):
        extra = extra.strip()
        if extra:
            try:
                images = images + extract_gallery_images(fetch_text(extra))
            except urllib.error.HTTPError as exc:
                print(f"  ::warning:: extra page HTTP {exc.code}: {extra}")

    if not images and not config.get("image_override"):
        raise RuntimeError(f"no gallery images: {page}")

    image_url = pick_image_url(
        images,
        config["size"],
        config["sku_hint"],
        config.get("image_override", ""),
    )
    row["image_source_url"] = image_url
    print(f"  image: {image_url.rsplit('/', 1)[-1]}")

    if force or not main_path.is_file():
        main_path.parent.mkdir(parents=True, exist_ok=True)
        main_path.write_bytes(to_jpeg(fetch_bytes(image_url)))
    if force or not thumb_path.is_file():
        make_thumbnail(main_path, thumb_path)

    row["image_main"] = str(main_path.relative_to(ROOT)).replace("\\", "/")
    row["image_thumbnail"] = str(thumb_path.relative_to(ROOT)).replace("\\", "/")
    print(f"  wrote {main_path.name} ({main_path.stat().st_size} bytes)")
    print(f"  wrote {thumb_path.name} ({thumb_path.stat().st_size} bytes)")
    sync_to_volusion(code)
    return row


def sync_to_volusion(code: str) -> None:
    """Copy catalog bed images into vspfiles/photos for Volusion PLP/PDP."""
    VOLUSION_PHOTOS.mkdir(parents=True, exist_ok=True)
    for suffix in ("-1.jpg", "-1T.jpg"):
        src = IMAGE_DIR / f"{code}{suffix}"
        if src.is_file():
            dest = VOLUSION_PHOTOS / src.name
            shutil.copy2(src, dest)
            print(f"  sync {dest.relative_to(ROOT)}")


def write_csv(rows: list[dict[str, str]]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "volusion_code",
        "collection",
        "finish",
        "size",
        "product_name",
        "steve_silver_sku",
        "steve_silver_url",
        "description",
        "techspecs",
        "related_pieces",
        "image_main",
        "image_thumbnail",
        "image_source_url",
    ]
    with CSV_PATH.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, str]]) -> None:
    try:
        from openpyxl import Workbook  # noqa: PLC0415
        from openpyxl.utils import get_column_letter  # noqa: PLC0415
    except ImportError:
        print("openpyxl not installed — CSV only")
        return

    xlsx_path = OUT_DIR / "steve_silver_beds.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Steve Silver Beds"
    fieldnames = [
        "volusion_code",
        "collection",
        "finish",
        "size",
        "product_name",
        "steve_silver_sku",
        "steve_silver_url",
        "description",
        "techspecs",
        "related_pieces",
        "image_main",
        "image_thumbnail",
        "image_source_url",
    ]
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(f, "") for f in fieldnames])
    for idx, field in enumerate(fieldnames, start=1):
        width = 18 if field != "techspecs" else 60
        ws.column_dimensions[get_column_letter(idx)].width = width
    wb.save(xlsx_path)
    print(f"Wrote {xlsx_path.relative_to(ROOT)}")


def sync_all_to_volusion() -> int:
    count = 0
    for path in sorted(IMAGE_DIR.glob("SS-*FB*-*.jpg")):
        dest = VOLUSION_PHOTOS / path.name
        shutil.copy2(path, dest)
        count += 1
    return count


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--force", action="store_true", help="Re-download images")
    parser.add_argument("--sync-only", action="store_true", help="Copy catalog images to vspfiles/photos only")
    parser.add_argument("--code", action="append", help="Only these Volusion codes")
    args = parser.parse_args()

    if args.sync_only:
        n = sync_all_to_volusion()
        print(f"Synced {n} bed photo(s) to {VOLUSION_PHOTOS.relative_to(ROOT)}/")
        return 0 if n else 1

    products = BED_PRODUCTS
    if args.code:
        products = [p for p in BED_PRODUCTS if p["volusion_code"] in args.code]

    rows: list[dict[str, str]] = []
    fail = 0
    for config in products:
        try:
            rows.append(process_product(config, args.force))
        except Exception as exc:  # noqa: BLE001
            fail += 1
            print(f"  ::error:: {config['volusion_code']}: {exc}")

    write_csv(rows)
    write_xlsx(rows)
    synced = sync_all_to_volusion()
    print(f"\nSynced {synced} bed photo(s) to {VOLUSION_PHOTOS.relative_to(ROOT)}/")
    print(f"\nWrote {CSV_PATH.relative_to(ROOT)} ({len(rows)} rows)")
    print(f"Images in {IMAGE_DIR.relative_to(ROOT)}/")
    print(f"Done: {len(rows)} ok, {fail} failed")
    return 0 if fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
